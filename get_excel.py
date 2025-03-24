import json

import openpyxl

from constants import JUDGES, VERSIONS, ANSWER_MODEL_PAIRS, TEMPERATURE, SPLIT_TYPE_TO_PATH, QUERY_TYPE_TO_PATH, \
    SPLIT_NUM_TO_PATH
from utils import extract_ans_rulellm


def consistent_judge(model1, model2, judger, split_type='equal', split_num=3,
                     output_dir=r'./llm_judge_repo_data/vicuna_bench/order_change_judgment', version='old',
                     query_type='pure', temperature=TEMPERATURE):
    result_dir = output_dir

    if split_type in SPLIT_TYPE_TO_PATH.keys():
        result_dir += SPLIT_TYPE_TO_PATH[split_type]
    elif query_type in QUERY_TYPE_TO_PATH.keys():
        result_dir += QUERY_TYPE_TO_PATH[query_type]

    if split_num in SPLIT_NUM_TO_PATH.keys():
        result_dir += SPLIT_NUM_TO_PATH[split_num]

    output_file = f"{result_dir}/{query_type}_temp{temperature}_{model1}_{model2}_{version}_{judger}.jsonl"

    consist_list, inconsist_list = [], []

    try:
        with open(output_file, 'r', encoding='utf-8') as f1:
            queried_str = f1.readlines()
            assert len(queried_str) % 2 == 0, "The number of samples is even"

            for i in range(len(queried_str) // 2):
                first_sample = json.loads(queried_str[2 * i])
                second_sample = json.loads(queried_str[2 * i + 1])

                order1_ans = first_sample['judge_answer']
                order2_ans = second_sample['judge_answer']
                question_id = first_sample['question_id']

                order1_ans_ext = extract_ans_rulellm(order1_ans, version)
                order2_ans_ext = extract_ans_rulellm(order2_ans, version)

                # Tie or not equal answer
                if order1_ans_ext == order2_ans_ext and order1_ans_ext == 3 or order1_ans_ext != order2_ans_ext:
                    consist_list.append(question_id)
                # Not equal answer
                elif order1_ans_ext == order2_ans_ext:
                    inconsist_list.append(question_id)

    except FileNotFoundError:
        print('No such file exists!', output_file)

    return consist_list, inconsist_list


def _get_res_ans_dict_basis(model_1, model_2, version):
    return {
        'model1': model_1,
        'model2': model_2,
        'version': version,
        'temperature': TEMPERATURE,
        'pure_incon_rate': -1,
        'eqaulsplit_incon_rate': -1,
        'IBMsplit_incon_rate': -1,
        'equalsplit_fix_rate': -1,
        'IBMsplit_fix_rate': -1,
        'EqualIBMsplit_fix_rate': -1,
        'EqualIBMsplit_incon_rate': -1
    }


def write_to_excel(res, judger='gpt-3.5-turbo'):
    if not res:
        raise ValueError("The result list is empty.")

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"

    headers = [
        "model1", "model2", "pure_incon_rate", "eqaulsplit_incon_rate",
        "IBMsplit_incon_rate", "EqualIBMsplit_incon_rate",
        "equalsplit_fix_rate", "IBMsplit_fix_rate", "EqualIBMsplit_fix_rate"
    ]

    # Write column headers
    for col, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col).value = header

    # Write data rows
    for row_idx, res_ans_dict in enumerate(res, start=2):
        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(row=row_idx, column=col_idx).value = res_ans_dict.get(header)

    # Generate filename based on judger, version, and temperature
    version = res[0].get('version', 'unknown')
    temperature = str(res[0].get('temperature', 'unknown'))
    filename = f'output_{judger}_version_{version}_temp_{temperature}.xlsx'

    workbook.save(filename=filename)


def main():
    for judge in JUDGES:
        for version in VERSIONS:  #
            judge_version_result = []

            for model_pair in ANSWER_MODEL_PAIRS:
                model_1 = model_pair[0]
                model_2 = model_pair[1]
                print(f"Model 1: {model_1}")
                print(f"Model 2: {model_2}")

                res_ans_dict = _get_res_ans_dict_basis(model_1, model_2, version)

                consistent_list_1, inconsistent_list_1 = consistent_judge(
                    model1=model_1, model2=model_2, judger=judge,
                    version=version
                )

                consistent_list_2, inconsistent_list_2 = consistent_judge(
                    model1=model_1, model2=model_2, judger=judge,
                    version=version, query_type='split'
                )

                consistent_list_3, inconsistent_list_3 = consistent_judge(
                    model1=model_1, model2=model_2, judger=judge,
                    split_type='IBM-pure', version=version, query_type='split'
                )

                eq_ibm_inconsistent_list = list(set(inconsistent_list_2).intersection(set(inconsistent_list_3)))
                eq_ibm_consistent_list = list(set(consistent_list_2).union(set(consistent_list_3)))

                if len(inconsistent_list_1) != 0 and len(inconsistent_list_2) != 0 and len(inconsistent_list_3) != 0:
                    fixed_list_1 = list(set(inconsistent_list_1).intersection(set(consistent_list_2)))
                    fixed_list_2 = list(set(inconsistent_list_1).intersection(set(consistent_list_3)))
                    fixed_list_3 = list(set(inconsistent_list_1).intersection(set(eq_ibm_consistent_list)))

                    res_ans_dict['pure_incon_rate'] = len(inconsistent_list_1) * 1.0 / 80
                    res_ans_dict['eqaulsplit_incon_rate'] = len(inconsistent_list_2) * 1.0 / 80
                    res_ans_dict['IBMsplit_incon_rate'] = len(inconsistent_list_3) * 1.0 / 80
                    res_ans_dict['EqualIBMsplit_incon_rate'] = len(eq_ibm_inconsistent_list) * 1.0 / 80
                    res_ans_dict['equalsplit_fix_rate'] = len(fixed_list_1) * 1.0 / len(inconsistent_list_1)
                    res_ans_dict['IBMsplit_fix_rate'] = len(fixed_list_2) * 1.0 / len(inconsistent_list_1)
                    res_ans_dict['EqualIBMsplit_fix_rate'] = len(fixed_list_3) * 1.0 / len(inconsistent_list_1)

                judge_version_result.append(res_ans_dict)

            write_to_excel(judge_version_result, judger=judge)


if __name__ == "__main__":
    main()
