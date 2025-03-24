import json 
import os
# from chatgpt import chat
import time
# from qwen import qwenchat
from chatglm import chatglmchat
# from minmax import call_minmax
# from exttool import qualitycheck
# from claude import call_claude
# from gpt4 import callgpt4
# from gpt4new import callgpt4new
# from IBM6 import IBM6
# from llama2 import call_llama
import openpyxl
from tqdm import tqdm

def split_ans(ans1,ans2,question_content, split_num=3 , split_type='equal',version='old',only_code=False, additional_info=False):
    template_init = f"\n[Question]\n{question_content}\n\n"
    names = ['A','B']
    split_code_success = False
    split_IBM_success = False
    if split_type == 'equal' and split_num>1 and not only_code: 
        ans1_split_len = len(ans1)//split_num
        ans2_split_len = len(ans2)//split_num
        for part_num in range(0, split_num):
            ans1_part = ans1[part_num * ans1_split_len:(part_num + 1) * ans1_split_len]
            ans2_part = ans2[part_num * ans2_split_len:(part_num + 1) * ans2_split_len]
            template_init += f"[The Start of Assistant {names[0]}'s Answer part{part_num + 1}]\n{ans1_part}\n\n[The End of Assistant {names[0]}'s Answer part{part_num + 1}]\n\n[The Start of Assistant {names[1]}'s Answer part{part_num + 1}]\n{ans2_part}\n\n[The End of Assistant {names[1]}'s Answer part{part_num + 1}]\n\n"
    elif split_type == 'IBM-pure':
        res, best_score, best_ans1, best_ans2 = IBM6(ans1=ans1,ans2=ans2,split_num=split_num,model_name='pure',length_penalty=True)
        if res: 
            for part_num in range(0, split_num):
                ans1_part = best_ans1[part_num]
                ans2_part = best_ans2[part_num]
                template_init += f"[The Start of Assistant {names[0]}'s Answer part{part_num + 1}]\n{ans1_part}\n\n[The End of Assistant {names[0]}'s Answer part{part_num + 1}]\n\n[The Start of Assistant {names[1]}'s Answer part{part_num + 1}]\n{ans2_part}\n\n[The End of Assistant {names[1]}'s Answer part{part_num + 1}]\n\n"
                split_IBM_success = True
        else:
            return split_ans(ans1,ans2,question_content, split_num=split_num , split_type='equal',version=version,only_code=only_code, additional_info=additional_info)
    elif split_type =='IBM-BERT':
        res, best_score, best_ans1, best_ans2 = IBM6(ans1=ans1,ans2=ans2,split_num=split_num,model_name='bert-base-nli-mean-tokens',length_penalty=True)
        if res: 
            for part_num in range(0, split_num):
                ans1_part = best_ans1[part_num]
                ans2_part = best_ans2[part_num]
                template_init += f"[The Start of Assistant {names[0]}'s Answer part{part_num + 1}]\n{ans1_part}\n\n[The End of Assistant {names[0]}'s Answer part{part_num + 1}]\n\n[The Start of Assistant {names[1]}'s Answer part{part_num + 1}]\n{ans2_part}\n\n[The End of Assistant {names[1]}'s Answer part{part_num + 1}]\n\n"
                split_IBM_success = True
        else:
            return split_ans(ans1,ans2,question_content, split_num=split_num , split_type='equal',version=version,only_code=only_code, additional_info=additional_info)



    if version == 'old': # score-based
        system_prompt = "We would like to request your feedback on the performance of two AI assistants in response to the user question displayed above.\nPlease rate the helpfulness, relevance, accuracy, level of details of their responses. Each assistant receives an overall score on a scale of 1 to 10, where a higher score indicates better overall performance.\nPlease first output a single line containing only two values indicating the scores for Assistant A and B, respectively. The two scores are separated by a space. In the subsequent line, please provide a comprehensive explanation of your evaluation, avoiding any potential bias and ensuring that the order in which the responses were presented does not affect your judgment.\n\n"
    elif version == 'new': # relation-based
        system_prompt = "Please act as an impartial judge and evaluate the quality of the responses provided by two AI assistants to the user question displayed below. You should choose the assistant that follows the user's instructions and answers the user's question better. Your evaluation should consider factors such as the helpfulness, relevance, accuracy, depth, creativity, and level of detail of their responses. Begin your evaluation by comparing the two responses and provide a short explanation. Avoid any positional biases and ensure that the order in which the responses were presented does not influence your decision. Do not allow the length of the responses to influence your evaluation. Do not favor certain names of the assistants. Be as objective as possible. After providing your explanation, output your final verdict by strictly following this format: \"[[A]]\" if assistant A is better, \"[[B]]\" if assistant B is better, and \"[[C]]\" for a tie."
    elif version =='likert': # likert-based
        system_prompt =f"We would like to request your feedback on the performance of two AI assistants in response to the user question displayed above.Please compare the helpfulness, relevance, accuracy, level of details of their responses. The rating should be from the set of 1, 2, 3, 4, 5, 6, or 7, where higher numbers indicated that Assistant {names[1]} was better than Assistant {names[0]}. Please first output a single line containing only one value indicating the preference between Assistant {names[0]} and {names[1]}. In the subsequent line, please provide a brief explanation of your evaluation, avoiding any potential bias and ensuring that the order in which the responses were presented does not affect your judgment."
    prompt = template_init + "[System]\n" + system_prompt 
    prompt = prompt.replace('\n\n','\n')
    return  prompt,split_code_success,split_IBM_success


def extract_ans(answer, version='old'):
    if answer is None:
        return -1
    if version == 'old':
        try:
            score1 = float(answer.split('\n')[0].split(' ')[0])
            score2 = float(answer.split('\n')[0].split(' ')[1])
            if score1 > score2:
                return 1
            elif score1 < score2:
                return 2
            else:
                return 3
        except:
            return -1
    elif version == 'new':
        if '[[A]]' in answer:
            return 1
        elif '[[B]]' in answer:
            return 2
        elif '[[C]]' in answer:
            return 3
        else:
            return -1
    elif version == 'likert':
        try:
            score = int(answer.split('\n')[0].strip())
            if score < 4:
                return 1
            elif score > 4:
                return 2
            elif score == 4:
                return 3
        except:
            return -1

def extract_ans_rulellm(answer, version='old'):
    if answer is None:
        return -1
    res = extract_ans(answer, version=version)
    if res != -1: 
        return res 
    prompt = 'You are given a paragraph of judgement. You are asked to tell me which AI assistant is better according to the judgement. '
    prompt += '[Judgement] '
    prompt += answer
    prompt += '\n return \"[[A]]\" if assistant A is better, return \"[[B]]\" if assistant B is better, and return \"[[C]]\" for a tie. Only return one string.'
    prompt += 'Only return one string in one line, do not explain, do not return two strings.'
    llmanswer = chatglmchat(question = prompt, temperature = 0)
    res = extract_ans(llmanswer, version='new')
    return res


def consistent_judge(model1='llama-13b', model2='vicuna-13b',split_type='equal',split_num='3', judger='gpt-3.5-turbo', question_dir=r'../llm_judge_repo_data/data/questions/question.jsonl', answer_dir=r'../llm_judge_repo_data/data/vicuna_bench/model_answer', output_dir=r'../llm_judge_repo_data/data/vicuna_bench/order_change_judgment',version='old', query_type='pure',temperature=0.7, only_code=False, additional_info = False):
    output_dir2 = output_dir
    if only_code: 
        output_dir2 = output_dir2 + '/only_code'
    if additional_info: 
        output_dir2 = output_dir2 + '/additional_info'
    if split_type =='IBM-BERT': 
        output_dir2 = output_dir2 + '/IBM6BERT'
    if split_type =='IBM-pure': 
        output_dir2 = output_dir2 + '/IBM6PURE'
    if split_type =='equal' and query_type == 'split': 
        output_dir2 = output_dir2 + '/EqualSplit'
    if split_type =='equal' and query_type == 'pure': 
        output_dir2 = output_dir2 + '/Pure'
    if split_num==2:
        output_dir2 += 'Split2'
    elif split_num==4:
        output_dir2 += 'Split4'
    output_dir = output_dir2
    IBM_num, total_num = 0,0
    output_file = output_dir + '/' + query_type +'_'+'temp'+ str(temperature)+ '_' + model1 + '_'+model2+'_'+ version + '_' + judger+'.jsonl'
    consist_list, inconsist_list = [],[]
    consist_num, inconsist_num, error_num = 0,0,0
    if not os.path.exists(output_file):
        print('no such file exists!', output_file)
        return [], []
    with open(output_file,'r', encoding='utf-8') as f1:
        queried_str = f1.readlines()
        
        assert len(queried_str)%2 == 0
        for i in range(len(queried_str)//2):
            order1_ans = json.loads(queried_str[2*i])['judge_answer'] 
            order2_ans = json.loads(queried_str[2*i+1])['judge_answer'] 
            question_id = json.loads(queried_str[2*i])['question_id']

            order1_ans_ext = extract_ans_rulellm(order1_ans,version)
            order2_ans_ext = extract_ans_rulellm(order2_ans,version) 
            if order1_ans_ext == -1 or order2_ans_ext == -1:
                error_num += 1
            elif order1_ans_ext == order2_ans_ext and order1_ans_ext==3:  
                consist_num += 1
                consist_list.append(question_id)
            elif order1_ans_ext != order2_ans_ext:  
                consist_num += 1
                consist_list.append(question_id)
            elif order1_ans_ext == order2_ans_ext:
                inconsist_num += 1
                inconsist_list.append(question_id)

    
    # print(' version {}, query_type {}, split_type {}, only_code {} consistent num: {}, inconsistent num: {}, error num: {}, IBM_success num {} inconsistent list: {}'.format(version, query_type, split_type, only_code ,consist_num,inconsist_num,error_num,IBM_num,inconsist_list))
    return consist_list, inconsist_list

def extract_gpt4_eval(quesiton_id, model_1, model_2,type='pair'):
    gpt4_ans_dir = r'../llm_judge_repo_data/data/vicuna_bench/model_judgment'
    if type == 'pair':
        gpt4_ans_file = gpt4_ans_dir + '/' + 'gpt-4_pair' + '.jsonl'
    with open(gpt4_ans_file,'r', encoding='utf-8') as f1:
        gpt4_ans = f1.readlines()
        for item in gpt4_ans:
            gpt4ans = json.loads(item)
            q_id = gpt4ans["question_id"]
            if q_id != quesiton_id:
                continue
            else:
                c_model_1 = gpt4ans["model_1"]
                c_model_2 = gpt4ans["model_2"]
                if c_model_1 == model_1 and c_model_2 == model_2:
                    g1_winner,g2_winner = gpt4ans['g1_winner'],gpt4ans['g2_winner']
                    if g1_winner == 'tie':
                        return 3
                    if g1_winner == g2_winner and g1_winner == 'model_1':
                        return 1
                    elif g1_winner == g2_winner and g1_winner == 'model_2':
                        return 2
                    elif g1_winner != g2_winner:
                        return 0
                    else:
                        print('possible error! please check the gpt4 answer file question id: ', quesiton_id)
                        return -1
                elif c_model_1 == model_2 and c_model_2 == model_1:
                    g1_winner,g2_winner = gpt4ans['g1_winner'],gpt4ans['g2_winner']
                    if g1_winner == 'tie':
                        return 3
                    if g1_winner == g2_winner and g1_winner == 'model_1':
                        return 2 
                    elif g1_winner == g2_winner and g1_winner == 'model_2':
                        return 1 
                    elif g1_winner != g2_winner:
                        return 0
                    else:
                        print('possible error! please check the gpt4 answer file question id: ', quesiton_id)
                        return -1

def after_order_gpt4_both(model1='llama-13b', model2='vicuna-13b',split_type='equal',split_num='3', judger='gpt-3.5-turbo', question_dir=r'../llm_judge_repo_data/data/questions/question.jsonl', answer_dir=r'../llm_judge_repo_data/data/vicuna_bench/model_answer', output_dir=r'../llm_judge_repo_data/data/vicuna_bench/order_change_judgment',version='old', query_type='pure',temperature=0.7, only_code=False, additional_info = False,fixed_list = []):
    output_dir2 = output_dir
    output_dir_still = output_dir
    if only_code: 
        output_dir2 = output_dir2 + '/only_code'
    if additional_info: 
        output_dir2 = output_dir2 + '/additional_info'
    if split_type =='IBM-BERT': 
        output_dir2 = output_dir2 + '/IBM6BERT'
    if split_type =='IBM-pure': 
        output_dir2 = output_dir2 + '/IBM6PURE'
    if split_type =='equal' and query_type == 'split': 
        output_dir2 = output_dir2 + '/EqualSplit'
    if split_type =='equal' and query_type == 'pure': 
        output_dir2 = output_dir2 + '/Pure'
    if split_num==2:
        output_dir2 += 'Split2'
    output_dir = output_dir2
    output_file = output_dir + '/' + query_type +'_'+'temp'+ str(temperature)+ '_' + model1 + '_'+model2+'_'+ version + '_' + judger+'.jsonl'
    if query_type == 'split':
        output_file = output_dir_still +'/EqualSplit' + '/' + query_type +'_'+'temp'+ str(temperature)+ '_' + model1 + '_'+model2+'_'+ version + '_' + judger+'.jsonl'
        output_file2 = output_dir_still + '/IBM6PURE' + '/' + query_type +'_'+'temp'+ str(temperature)+ '_' + model1 + '_'+model2+'_'+ version + '_' + judger+'.jsonl'

    consist_list, inconsist_list = [],[]
    consist_num, inconsist_num, error_num = 0,0,0
    orderCon_gpt4Con_same, orderCon_gpt4Con_insame, orderCon_gpt4Incon, orderIncon_gpt4Con, orderIncon_gpt4Incon = 0,0,0,0,0
    gpt4consist_list, gpt4inconsist_list = [],[]
    split_fix_work_list = [] 
    split_fix_work_gpt4fail_list = [] 
    split_fix_notwork_list = [] 
    if not os.path.exists(output_file):
        print('no such file exists!', output_file)
        return False
    with open(output_file,'r', encoding='utf-8') as f1:
        queried_str = f1.readlines()
        assert len(queried_str)%2 == 0
    if query_type == 'split':
        with open(output_file2,'r', encoding='utf-8') as f2: 
            queried_str2 = f2.readlines()
            assert len(queried_str2)%2 == 0


    answer_dict = {} 
    for question_id in range(1,81):
        gpt4_ans = extract_gpt4_eval(question_id,model1,model2)
        answer_dict[question_id] = {'gpt4': gpt4_ans} 
    for i in range(len(queried_str)//2):
        order1_ans = json.loads(queried_str[2*i])['judge_answer'] 
        order2_ans = json.loads(queried_str[2*i+1])['judge_answer'] 
        question_id = json.loads(queried_str[2*i])['question_id'] 
        order1_ans_ext = extract_ans_rulellm(order1_ans,version)
        order2_ans_ext = extract_ans_rulellm(order2_ans,version) 
        if order1_ans_ext == -1 or order2_ans_ext == -1:
            answer_dict[question_id]['queried_str'] = -1
        if (order1_ans_ext != order2_ans_ext) or (order1_ans_ext==3 and order2_ans_ext==3):
            answer_dict[question_id]['queried_str'] =  order1_ans_ext
        elif order1_ans_ext == order2_ans_ext: 
            answer_dict[question_id]['queried_str'] =  0
    if query_type == 'split':
        for i in range(len(queried_str2)//2):
            order1_ans = json.loads(queried_str2[2*i])['judge_answer'] 
            order2_ans = json.loads(queried_str2[2*i+1])['judge_answer'] 
            question_id = json.loads(queried_str2[2*i])['question_id'] 
            order1_ans_ext = extract_ans_rulellm(order1_ans,version)
            order2_ans_ext = extract_ans_rulellm(order2_ans,version) 
            if order1_ans_ext == -1 or order2_ans_ext == -1:
                answer_dict[question_id]['queried_str2'] = -1
            if (order1_ans_ext != order2_ans_ext) or (order1_ans_ext==3 and order2_ans_ext==3):
                answer_dict[question_id]['queried_str2'] =  order1_ans_ext
            elif order1_ans_ext == order2_ans_ext: 
                answer_dict[question_id]['queried_str2'] =  0

    gpt4same = 0
    for question_id, value in answer_dict.items(): 
        if value['gpt4'] not in [-1,0]: 
            gpt4same += 1
            consistent_flag = False 
            if 'queried_str' in value: 
                if value['queried_str'] == value['gpt4']:
                    consistent_flag = True
            if 'queried_str2' in value: 
                if value['queried_str2'] == value['gpt4']:
                    consistent_flag = True
            if consistent_flag:
                orderCon_gpt4Con_same += 1
    print('orderCon_gpt4Con_same {}, gpt4same {}'.format(orderCon_gpt4Con_same,gpt4same))
    return orderCon_gpt4Con_same, gpt4same


def main():
    temperature = 0
    for judger in ['gpt-3.5-turbo']: # 'gpt-3.5-turbo'  'chatglm',  'qwen', 'claude2',  'gpt4newtx'
        for version in ['new']:  # ['new','old','likert']
            models = [['gpt-3.5-turbo', 'claude-v1'],['llama-13b', 'vicuna-13b'],['alpaca-13b', 'vicuna-13b'], 
                ['gpt-3.5-turbo', 'gpt-4'], 
                ['gpt-4', 'claude-v1'], 
                ['vicuna-13b', 'vicuna-7b'], 
                ['vicuna-7b', 'alpaca-13b'], 
                ['gpt-4', 'vicuna-13b']] 

            models = [['gpt-3.5-turbo', 'claude-v1']]
            res = []

            for model_pair in models:

                model1 = model_pair[0]
                model2 = model_pair[1]
                print(f"Model 1: {model1}")
                print(f"Model 2: {model2}")

                res_ans_dict = {'model1':model1,'model2':model2,'version':version,'temperature':temperature,'pure_incon_rate':-1,'eqaulsplit_incon_rate':-1,'IBMsplit_incon_rate':-1,'equalsplit_fix_rate':-1, 'IBMsplit_fix_rate':-1, 'EqualIBMsplit_fix_rate':-1, 'EqualIBMsplit_incon_rate': -1}


                consist_list1, inconsist_list1 = consistent_judge(model1=model1, model2=model2,judger=judger,version=version,query_type='pure',temperature=temperature)
                consist_list2, inconsist_list2 = consistent_judge(model1=model1, model2=model2,judger=judger,version=version,query_type='split',temperature=temperature)
                consist_list6, inconsist_list6 = consistent_judge(model1=model1, model2=model2,judger=judger,split_type='IBM-pure', split_num=3, version=version,query_type='split',temperature=temperature,only_code=False,additional_info=False)
                
                Eq_IBM_incon_list = list(set(inconsist_list2).intersection(set(inconsist_list6)))
                Eq_IBM_con_list = list(set(consist_list2).union(set(consist_list6)))

                if len(inconsist_list1) == 0  or len(inconsist_list2) == 0 or len(inconsist_list6) == 0:
                    res.append(res_ans_dict)
                    continue


                fixed_list1 = list(set(inconsist_list1).intersection(set(consist_list2)))
                fixed_list2 = list(set(inconsist_list1).intersection(set(consist_list6)))
                fixed_list3 = list(set(inconsist_list1).intersection(set(Eq_IBM_con_list)))

                if len(inconsist_list1) != 0:
                    res_ans_dict['pure_incon_rate'] = len(inconsist_list1)*1.0 / 80
                    res_ans_dict['eqaulsplit_incon_rate'] = len(inconsist_list2)*1.0 / 80
                    res_ans_dict['IBMsplit_incon_rate'] = len(inconsist_list6)*1.0 / 80
                    res_ans_dict['EqualIBMsplit_incon_rate'] = len(Eq_IBM_incon_list)*1.0 / 80
                    res_ans_dict['equalsplit_fix_rate'] = len(fixed_list1)*1.0/len(inconsist_list1) 
                    res_ans_dict['IBMsplit_fix_rate'] = len(fixed_list2)*1.0/len(inconsist_list1)
                    res_ans_dict['EqualIBMsplit_fix_rate'] = len(fixed_list3)*1.0/len(inconsist_list1)
                res.append(res_ans_dict)

            write_to_excel(res,judger=judger)
        

def write_to_excel(res,judger='gpt-3.5-turbo'):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"

    worksheet.cell(row=1, column=1).value = "model1"
    worksheet.cell(row=1, column=2).value = "model2"
    worksheet.cell(row=1, column=3).value = "pure_incon_rate"
    worksheet.cell(row=1, column=4).value = "eqaulsplit_incon_rate"
    worksheet.cell(row=1, column=5).value = "IBMsplit_incon_rate"
    worksheet.cell(row=1, column=6).value = "EqualIBMsplit_incon_rate"
    worksheet.cell(row=1, column=7).value = "equalsplit_fix_rate"
    worksheet.cell(row=1, column=8).value = "IBMsplit_fix_rate"
    worksheet.cell(row=1, column=9).value = "EqualIBMsplit_fix_rate"
    row = 2
    version,temperature = res[0]['version'], str(res[0]['temperature'])
    for res_ans_dict in res:
        worksheet.cell(row=row, column=1).value = res_ans_dict['model1']
        worksheet.cell(row=row, column=2).value = res_ans_dict['model2']
        worksheet.cell(row=row, column=3).value = res_ans_dict['pure_incon_rate']
        worksheet.cell(row=row, column=4).value = res_ans_dict['eqaulsplit_incon_rate']
        worksheet.cell(row=row, column=5).value = res_ans_dict['IBMsplit_incon_rate']
        worksheet.cell(row=row, column=6).value = res_ans_dict['EqualIBMsplit_incon_rate']
        worksheet.cell(row=row, column=7).value = res_ans_dict['equalsplit_fix_rate']
        worksheet.cell(row=row, column=8).value = res_ans_dict['IBMsplit_fix_rate']
        worksheet.cell(row=row, column=9).value = res_ans_dict['EqualIBMsplit_fix_rate']
        row += 1
    filename = 'output_' + judger +'_version_' +version +'_temp_' + temperature+'.xlsx'
    workbook.save(filename=filename)

def RQ_ablation():
    temperature = 0
    model1='gpt-3.5-turbo'
    model2='claude-v1'
    judger= 'llama2-70b'#  'gpt-3.5-turbo'  'chatglm',  'qwen', 'minmax', 'claude2', 'gpt4', 'llama2-70b', 'gpt4new'
    version = 'new' 
    orderCon_gpt4Con_origin, gptsame_origin = 0,0
    orderCon_gpt4Con_fix, gptsame_fix = 0,0

    for judger in ['gpt-3.5-turbo']:
        # for version in ['new', 'old', 'likert']:
        for version in ['new']:
            models = [['gpt-3.5-turbo', 'claude-v1'],['llama-13b', 'vicuna-13b'],['alpaca-13b', 'vicuna-13b'], 
                ['gpt-3.5-turbo', 'gpt-4'], 
                ['gpt-4', 'claude-v1'], 
                ['vicuna-13b', 'vicuna-7b'], 
                ['vicuna-7b', 'alpaca-13b'], 
                ['gpt-4', 'vicuna-13b']]

            for model_pair in models:
                model1 = model_pair[0]
                model2 = model_pair[1]
                print('judge is ', judger)
                print(f"Model 1: {model1}")
                print(f"Model 2: {model2}")


                orderCon_gpt4Con_same1, gpt4_same1 = after_order_gpt4_both(model1=model1, model2=model2,judger=judger,version=version,query_type='pure',temperature=temperature)
                orderCon_gpt4Con_same2, gpt4_same2 = after_order_gpt4_both(model1=model1, model2=model2,judger=judger,version=version,query_type='split',split_type='IBM-pure',temperature=temperature)

                orderCon_gpt4Con_origin += orderCon_gpt4Con_same1 
                gptsame_origin += gpt4_same1
                orderCon_gpt4Con_fix+=orderCon_gpt4Con_same2
                gptsame_fix += gpt4_same2
        

        print('origin AR rate is {}'.format(round(orderCon_gpt4Con_origin*100.0/gptsame_origin,4)))
        print('fix AR rate is {}'.format(round(orderCon_gpt4Con_fix*100.0/gptsame_fix,4)))


if __name__ == "__main__":
    main()
    # RQ_ablation()