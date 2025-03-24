from openpyxl import load_workbook
import os
import json
import matplotlib.pyplot as plt
import numpy as np
import matplotlib.patches as mpatches
import argparse


def readnum(judge='gpt-3.5-turbo',version='likert',main_result=True):
    basedir = './'
    name = 'output_{}_version_{}_temp_0.xlsx'.format(judge,version)
    filename = basedir + name
    wb = load_workbook(filename)
    sheets = wb.sheetnames
    sheet1 = wb[sheets[0]]
    pure_incon_rate = []
    EqualIBMsplit_incon_rate = []
    EqualIBMsplit_fix_rate =[]
    eqaulsplit_incon_rate = []
    IBMsplit_incon_rate = []
    equalsplit_fix_rate = []
    IBMsplit_fix_rate = []

    pure_incon_rate_column = sheet1['C']
    eqaulsplit_incon_rate_column = sheet1['D']
    IBMsplit_incon_rate_column = sheet1['E']
    EqualIBMsplit_incon_rate_column = sheet1['F']
    equalsplit_fix_rate_column = sheet1['G']
    IBMsplit_fix_rate_column = sheet1['H']
    EqualIBMsplit_fix_rate_column = sheet1['I']
    for x in range(1,len(pure_incon_rate_column)):
        if pure_incon_rate_column[x].value != -1:
            pure_incon_rate.append(pure_incon_rate_column[x].value) 
        if EqualIBMsplit_incon_rate_column[x].value != -1:
            EqualIBMsplit_incon_rate.append(EqualIBMsplit_incon_rate_column[x].value)
        if EqualIBMsplit_fix_rate_column[x].value != -1:
            EqualIBMsplit_fix_rate.append(EqualIBMsplit_fix_rate_column[x].value)
        if eqaulsplit_incon_rate_column[x].value != -1:
            eqaulsplit_incon_rate.append(eqaulsplit_incon_rate_column[x].value)
        if IBMsplit_incon_rate_column[x].value != -1:
            IBMsplit_incon_rate.append(IBMsplit_incon_rate_column[x].value)
        if equalsplit_fix_rate_column[x].value != -1:
            equalsplit_fix_rate.append(equalsplit_fix_rate_column[x].value)
        if IBMsplit_fix_rate_column[x].value != -1:
            IBMsplit_fix_rate.append(IBMsplit_fix_rate_column[x].value)
    
    gap = sum(pure_incon_rate)/len(pure_incon_rate) - sum(EqualIBMsplit_incon_rate)/len(EqualIBMsplit_incon_rate)
    if main_result:
        return sum(pure_incon_rate)/len(pure_incon_rate),sum(EqualIBMsplit_incon_rate)/len(EqualIBMsplit_incon_rate),gap,sum(EqualIBMsplit_fix_rate)/len(EqualIBMsplit_fix_rate)

    return sum(pure_incon_rate)/len(pure_incon_rate),sum(EqualIBMsplit_incon_rate)/len(EqualIBMsplit_incon_rate),gap,sum(EqualIBMsplit_fix_rate)/len(EqualIBMsplit_fix_rate), sum(eqaulsplit_incon_rate)/len(eqaulsplit_incon_rate),sum(IBMsplit_incon_rate)/len(IBMsplit_incon_rate),sum(equalsplit_fix_rate)/len(equalsplit_fix_rate),sum(IBMsplit_fix_rate)/len(IBMsplit_fix_rate)


def main_result_consistent(versions=['old','new','likert'], judges=['gpt4newtx','gpt-3.5-turbo','qwen','chatglm','claude2']):
    upnums = []
    for version in versions:
        for judge in judges:
            print('judge is {}, version is {} ######old is scorebased, new is relation based'.format(judge,version))
            pure_incon_rate,EqualIBMsplit_incon_rate,gap,EqualIBMsplit_fix_rate = readnum(judge,version)
            pure_incon_rate = round(pure_incon_rate*100,2)
            EqualIBMsplit_incon_rate = round(EqualIBMsplit_incon_rate*100,2)
            gap = round(gap*100,2)
            EqualIBMsplit_fix_rate = round(EqualIBMsplit_fix_rate*100,2)
            pure_con_rate = 100 - pure_incon_rate
            EqualIBMsplit_con_rate = 100 - EqualIBMsplit_incon_rate
            up = (EqualIBMsplit_con_rate/pure_con_rate)*100 -100
            upnum = round(up,2)
            print('pure_con_rate is {}, EqualIBMsplit_con_rate is {}, rise is $\\uparrow${}\\%, EqualIBMsplit_fix_rate is {}'.format(pure_con_rate,EqualIBMsplit_con_rate, upnum , EqualIBMsplit_fix_rate))
            upnums.append(upnum)
            print('-'*10)
    print('upnums is {}, average is {}'.format(upnums, sum(upnums)/len(upnums)))


def draw_ablation_pic_way2(draw='incon_rate',judges=['claude2','qwen','chatglm','gpt-3.5-turbo','gpt4newtx']):
    i = 0
    fig, ax = plt.subplots(figsize=(7,2))
    hatches = ['', '\\\\\\', '...', '\\']
    colors = ['#f0f7ec','#f4b384','#b6df9a','#77bbec','#f78280']
    colors = ['#77bbec','#f4b384','#b6df9a','#77bbec','#f78280']
    myfontsize = 8

    edgecolors = ['black','black','black','black']
    for version in ['old','new','likert']:
        x_axis = judges
        print('version is {}'.format(version))

        pure_incon_rate_list,EqualIBMsplit_incon_rate_list,EqualIBMsplit_fix_rate_list = [],[],[]
        eqaulsplit_incon_rate_list,IBMsplit_incon_rate_list,equalsplit_fix_rate_list,IBMsplit_fix_rate_list = [],[],[],[]
        for judge in x_axis:
            pure_incon_rate,EqualIBMsplit_incon_rate,gap,EqualIBMsplit_fix_rate,eqaulsplit_incon_rate,IBMsplit_incon_rate,equalsplit_fix_rate,IBMsplit_fix_rate = readnum(judge,version,main_result=False)
            pure_incon_rate = round(pure_incon_rate*100,2)
            EqualIBMsplit_incon_rate = round(EqualIBMsplit_incon_rate*100,2)
            EqualIBMsplit_fix_rate = round(EqualIBMsplit_fix_rate*100,2)
            eqaulsplit_incon_rate = round(eqaulsplit_incon_rate*100,2)
            IBMsplit_incon_rate = round(IBMsplit_incon_rate*100,2)
            equalsplit_fix_rate = round(equalsplit_fix_rate*100,2)
            IBMsplit_fix_rate = round(IBMsplit_fix_rate*100,2)
            pure_incon_rate_list.append(pure_incon_rate)
            EqualIBMsplit_incon_rate_list.append(EqualIBMsplit_incon_rate)
            EqualIBMsplit_fix_rate_list.append(EqualIBMsplit_fix_rate)
            eqaulsplit_incon_rate_list.append(eqaulsplit_incon_rate)
            IBMsplit_incon_rate_list.append(IBMsplit_incon_rate)
            equalsplit_fix_rate_list.append(equalsplit_fix_rate)
            IBMsplit_fix_rate_list.append(IBMsplit_fix_rate)

        x = np.arange(len(x_axis))
        width = 0.08
        print('hatch is {}'.format(hatches[i]))
        if draw == 'incon_rate':
            ax.bar(x - 2*width + i*4*width,pure_incon_rate_list, width=width,label='Original',hatch=hatches[i],edgecolor=edgecolors[0],color=colors[i])
            ax.bar(x - width + i*4*width, EqualIBMsplit_incon_rate_list, width=width, label='PORTIA',hatch=hatches[i],edgecolor=edgecolors[0],color=colors[i])
            ax.bar(x + i*4*width, eqaulsplit_incon_rate_list, width=width, label='PORTIA w/o SA',hatch=hatches[i],edgecolor=edgecolors[0],color=colors[i]) # semantic alignment
            ax.bar(x + width + i*4*width, IBMsplit_incon_rate_list, width=width, label='PORTIA w/o LA',hatch=hatches[i],edgecolor=edgecolors[0],color=colors[i]) # length alignment
            ax.axvline(x=x + 1.5*width + i*4*width, ls='--',c='black')
        elif draw == 'fix_rate':
            ax.bar(x-width + i*4*width, EqualIBMsplit_fix_rate_list, width=width, label='PORTIA',hatch=hatches[i],edgecolor=edgecolors[0],color=colors[0])
            ax.bar(x + i*4*width, equalsplit_fix_rate_list, width=width, label='PORTIA w/o SA',hatch=hatches[i],edgecolor=edgecolors[1],color=colors[1]) # semantic alignment
            ax.bar(x+width + i*4*width, IBMsplit_fix_rate_list, width=width, label='PORTIA w/o LA',hatch=hatches[i],edgecolor=edgecolors[2],color=colors[2]) # length alignment
        i += 1
        ax.axvline(x[i]-2.25*width, ls='--',c='#7c7c7c')

    i += 1
    ax.axvline(x[i]-2.25*width, ls='--',c='#7c7c7c')
    ax.set_xticks(x+4*width)
    ax.set_xticklabels(['Claude2','Qwen','Chatglm2','GPT-3.5','GPT-4'],fontsize=myfontsize)
    if draw == 'incon_rate':
        plt.ylabel('Inconclusive (%)', fontsize=myfontsize)
    elif draw == 'fix_rate':
        plt.ylabel('Fixed Coverage (%)', fontsize=myfontsize)
    
    rect_handle1 = plt.Rectangle((0, 0), 1, 1, edgecolor=edgecolors[0], facecolor=colors[0]) 
    rect_handle2 = plt.Rectangle((0, 0), 1, 1, edgecolor=edgecolors[0], facecolor=colors[1]) 
    rect_handle3 = plt.Rectangle((0, 0), 1, 1, edgecolor=edgecolors[0], facecolor=colors[2]) 
    rect_handle4 = plt.Rectangle((0, 0), 1, 1, hatch=hatches[0],edgecolor='black',facecolor='white') # 
    rect_handle5 = plt.Rectangle((0, 0), 1, 1, hatch=hatches[1],edgecolor='black',facecolor='white') # 
    rect_handle6 = plt.Rectangle((0, 0), 1, 1, hatch=hatches[2],edgecolor='black',facecolor='white') # 


    rect_handles = [rect_handle1, rect_handle2, rect_handle3, rect_handle4, rect_handle5, rect_handle6]
    mylabels = ['PORTIA','PORTIA w/o SA','PORTIA w/o LA','score-based','relation-based','likert-based']
    ax.legend(handles=rect_handles, labels=mylabels,loc='upper center', prop={'size': 6}, ncol=2)

    savename = r'./ablation-gpt4.pdf'
    plt.savefig(savename, bbox_inches='tight')

main_result_consistent()
draw_ablation_pic_way2(draw='fix_rate')


parser = argparse.ArgumentParser()
parser.add_argument("--RQ1", help="table for RQ1", action="store_true")
parser.add_argument("--RQ2", help="fig for Ablation study", action="store_true")

args = parser.parse_args()

if args.RQ1:
    main_result_consistent()
elif args.RQ2:
    draw_ablation_pic_way2(draw='fix_rate')
else:
    print("PLEASE USE --RQ1 OR --RQ2")
